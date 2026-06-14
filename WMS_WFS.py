import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import os
import webbrowser
import requests
import json
import folium
import geopandas as gpd
import xml.etree.ElementTree as ET
import time  
from owslib.wms import WebMapService
from owslib.wfs import WebFeatureService
from datetime import datetime
import uuid 
import pandas as pd

# Importación para manejo nativo y estético de Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ReportLab imports unificados
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfgen import canvas

# --- VARIABLES GLOBALES ---
NODOS_CATALOGO = []
wms_cached_layers = []
wfs_cached_layers = []
WMS_SERVER_METADATA = {}
WFS_SERVER_METADATA = {}
ARCHIVOS_TEMPORALES_MAPAS = []  

# --- FUNCIÓN PARA CALIFICAR LA LATENCIA ---
def calificar_latencia(ms):
    if ms < 200:
        return f"{ms:.0f} ms (Muy buena)"
    elif ms < 500:
        return f"{ms:.0f} ms (Buena)"
    elif ms < 1500:
        return f"{ms:.0f} ms (Regular)"
    else:
        return f"{ms:.0f} ms (Insuficiente)"

# --- LIMPIEZA DE ARCHIVOS TEMPORALES ---
def limpiar_archivos_temporales():
    archivos_temp = ["mapa_wms_visual.html", "mapa_wfs_auditoria.html"] + ARCHIVOS_TEMPORALES_MAPAS
    for archivo in archivos_temp:
        if os.path.exists(archivo):
            try:
                os.remove(archivo)
            except Exception as e:
                print(f"No se pudo eliminar {archivo}: {e}")

# --- CARGA DEL CATÁLOGO LOCAL ---
def cargar_catalogo_local():
    global NODOS_CATALOGO
    archivo_json = "catalogo.json"
    try:
        if os.path.exists(archivo_json):
            with open(archivo_json, 'r', encoding='utf-8') as f:
                NODOS_CATALOGO = json.load(f)
        else:
            status_label.config(text=f"Error: No se encontró '{archivo_json}'.", foreground="red")
            return
        filtrar_lista_catalogo()
    except Exception as e:
        status_label.config(text=f"Error al leer la base de datos local: {e}", foreground="red")

def filtrar_lista_catalogo(event=None):
    criterio = idera_search_entry.get().lower().strip()
    for item in idera_tree.get_children():
        idera_tree.delete(item)
        
    contador_visibles = 0
    for idx, nodo in enumerate(NODOS_CATALOGO):
        org = nodo["Organismo"]
        wms = nodo["WMS"]
        wfs = nodo["WFS"]
        
        if criterio in org.lower() or criterio in wms.lower() or criterio in wfs.lower():
            idera_tree.insert("", "end", iid=f"nodo_{idx}", values=(org, wms, wfs))
            contador_visibles += 1
            
    if criterio:
        status_label.config(text=f"Buscador: {contador_visibles} coincidencias.", foreground="blue")
    else:
        status_label.config(text=f"Catálogo: {len(NODOS_CATALOGO)} organismos disponibles.", foreground="green")

def inyectar_nodo_catalogo(tipo):
    selected = idera_tree.selection()
    if not selected:
        status_label.config(text="Por favor, seleccione un organismo de la lista.", foreground="red")
        return
    
    idx = int(selected[0].replace("nodo_", ""))
    nodo = NODOS_CATALOGO[idx]
    
    if tipo == 'WMS' and nodo["WMS"]:
        wms_url_entry.delete(0, tk.END)
        wms_url_entry.insert(0, nodo["WMS"])
        tab_control.select(wms_tab)  
        cargar_lista_wms()           
    elif tipo == 'WFS' and nodo["WFS"]:
        wfs_url_entry.delete(0, tk.END)
        wfs_url_entry.insert(0, nodo["WFS"])
        tab_control.select(wfs_tab)  
        cargar_lista_wfs()           
    else:
        status_label.config(text=f"Este organismo no dispone de endpoint {tipo}.", foreground="red")

# --- CONEXIÓN Y CARGA DE SERVICIOS ---
def cargar_lista_wms():
    global wms_cached_layers, WMS_SERVER_METADATA
    wms_url = wms_url_entry.get().strip()
    if not wms_url:
        status_label.config(text="Por favor, ingrese una URL WMS.", foreground="red")
        return
    
    for item in wms_tree.get_children():
        wms_tree.delete(item)
    wms_cached_layers = []
    WMS_SERVER_METADATA = {} 
    status_label.config(text="Conectando al servidor WMS...", foreground="blue")
    root.update_idletasks()

    try:
        inicio = time.time()  
        wms = WebMapService(wms_url, timeout=10)
        latencia_ms = (time.time() - inicio) * 1000 
        calificacion = calificar_latencia(latencia_ms)

        WMS_SERVER_METADATA = {
            'Formatos': getattr(wms, 'getmap_format_options', ["image/png", "image/jpeg"]),
            'Restricciones': getattr(wms.identification, 'accessconstraints', "Ninguna declarada"),
            'Tarifas': getattr(wms.identification, 'fees', "Gratuito / No declarado"),
            'Versión': wms.version,
            'Latencia': calificacion
        }
        
        for layer_name in wms.contents:
            layer = wms[layer_name]
            crs_list = getattr(layer, 'crsOptions', [])
            crs_string = ", ".join(list(crs_list)[:4]) if crs_list else "Declarados por el servidor"

            wms_cached_layers.append({
                'Checked': False,
                'Nombre': layer_name,
                'Título': layer.title or 'Sin Título',
                'Resumen': layer.abstract or 'Sin Resumen',
                'CRS': crs_string
            })
            
        actualizar_tabla_wms_visual()
        status_label.config(text=f"WMS Online - {len(wms_cached_layers)} capas. Latencia: {calificacion}", foreground="green")
    except requests.exceptions.ConnectionError:
        status_label.config(text="Error: El servidor WMS no responde. Latencia: Insuficiente", foreground="red")
    except Exception as e:
        status_label.config(text=f"Servidor WMS inalcanzable: {e}", foreground="red")


def cargar_lista_wfs_async():
    global wfs_cached_layers, WFS_SERVER_METADATA
    wfs_url = wfs_url_entry.get().split('?')[0].strip()

    if not wfs_url:
        status_label.config(text="Por favor, ingrese una URL WFS.", foreground="red")
        return

    for item in wfs_tree.get_children():
        wfs_tree.delete(item)

    wfs_cached_layers = []
    WFS_SERVER_METADATA = {}
    status_label.config(text="Conectando al servidor WFS...", foreground="blue")
    root.update_idletasks()

    try:
        wfs = None
        version_usada = None
        inicio = time.time() 

        for version in ['2.0.0', '1.1.0', '1.0.0']:
            try:
                wfs = WebFeatureService(wfs_url, version=version, timeout=15)
                version_usada = version
                break
            except Exception:
                continue

        if wfs is None:
            raise Exception("No fue posible conectar en ninguna versión WFS.")

        latencia_ms = (time.time() - inicio) * 1000
        calificacion = calificar_latencia(latencia_ms)

        try:
            operacion = wfs.getOperationByName('GetFeature')
            formatos_lista = getattr(operacion, 'formatOptions', [])
            if not formatos_lista and hasattr(operacion, 'parameters'):
                formatos_lista = operacion.parameters.get('outputFormat', {}).get('values', ['GML'])
        except Exception:
            formatos_lista = ['application/json', 'GeoJSON', 'GML']

        WFS_SERVER_METADATA = {
            'Version': version_usada,
            'Formatos': formatos_lista,
            'Restricciones': getattr(wfs.identification, 'accessconstraints', "Ninguna"),
            'Tarifas': getattr(wfs.identification, 'fees', "No declaradas"),
            'Titulo_Servidor': getattr(wfs.identification, 'title', "Servidor WFS"),
            'Latencia': calificacion
        }

        contenidos = list(wfs.contents.keys())
        for layer_name in contenidos:
            try:
                capa = wfs[layer_name]
                wfs_cached_layers.append({
                    'Checked': False,
                    'Nombre': layer_name,
                    'Título': capa.title or 'Sin Título',
                    'Resumen': capa.abstract or 'Sin Resumen',
                    'Registros': "Pendiente"
                })
            except Exception:
                continue

        actualizar_tabla_wfs_visual()
        status_label.config(text=f"WFS Online - {len(wfs_cached_layers)} capas. Latencia: {calificacion}", foreground="green")
    except requests.exceptions.ConnectionError:
        status_label.config(text="Error WFS: Servidor Caído. Latencia: Insuficiente", foreground="red")
    except Exception as e:
        status_label.config(text=f"Servidor WFS inalcanzable: {e}", foreground="red")

def cargar_lista_wfs():
    threading.Thread(target=cargar_lista_wfs_async, daemon=True).start()

# --- CONTROL DE SELECCIÓN DE TABLAS ---
def actualizar_tabla_wms_visual():
    for item in wms_tree.get_children():
        wms_tree.delete(item)
    for idx, layer in enumerate(wms_cached_layers):
        checkbox = "[X]" if layer.get('Checked', False) else "[  ]"
        wms_tree.insert("", "end", iid=f"wms_{idx}", values=(checkbox, layer['Nombre'], layer['Título']))

def alternar_seleccion_item_wms(event=None):
    sel = wms_tree.selection()
    if not sel: return
    idx = int(sel[0].replace("wms_",""))
    wms_cached_layers[idx]['Checked'] = not wms_cached_layers[idx].get('Checked', False)
    actualizar_tabla_wms_visual()

def cambiar_estado_todas_wms(estado):
    for l in wms_cached_layers: l['Checked'] = estado
    actualizar_tabla_wms_visual()

def actualizar_tabla_wfs_visual():
    for item in wfs_tree.get_children():
        wfs_tree.delete(item)
    for idx, layer in enumerate(wfs_cached_layers):
        checkbox = "[X]" if layer['Checked'] else "[  ]"
        reg_info = f"({layer['Registros']})" if layer['Registros'] == "Pendiente" else f"({layer['Registros']} reg.)"
        wfs_tree.insert("", "end", iid=str(idx), values=(checkbox, layer['Nombre'], f"{reg_info} {layer['Título']}"))

def alternar_seleccion_item(event=None):
    selected_item = wfs_tree.selection()
    if not selected_item: return
    idx = int(selected_item[0])
    wfs_cached_layers[idx]['Checked'] = not wfs_cached_layers[idx]['Checked']
    actualizar_tabla_wfs_visual()

def cambiar_estado_todas_wfs(estado):
    for layer in wfs_cached_layers: layer['Checked'] = estado
    actualizar_tabla_wfs_visual()

def obtener_hits_capa(url_wfs, layer_name, version_wfs):
    try:
        url_wfs_limpia = url_wfs.split('?')[0]
        params = {
            'service': 'WFS',
            'version': version_wfs,
            'request': 'GetFeature',
            'resultType': 'hits'
        }
        if str(version_wfs).startswith("2"):
            params['typeNames'] = layer_name
        else:
            params['typeName'] = layer_name

        res = requests.get(url_wfs_limpia, params=params, timeout=12)
        if res.status_code == 200:
            xml_root = ET.fromstring(res.content)
            for attr_key, attr_val in xml_root.attrib.items():
                if 'numberMatched' in attr_key or 'numberOfFeatures' in attr_key:
                    return int(attr_val)
    except Exception:
        pass
    return 0

# --- MÓDULO DE AUDITORÍA AVANZADA EN TABLA DE ATRIBUTOS ---
def ver_tabla_atributos_wfs():
    capas_tildadas = [c for c in wfs_cached_layers if c['Checked']]
    if not capas_tildadas:
        messagebox.showwarning("Atención", "Por favor, tilde [X] al menos una capa de la lista para auditar sus atributos.")
        return
    
    wfs_url = wfs_url_entry.get().split('?')[0].strip()
    
    ventana_pregunta = tk.Toplevel(root)
    ventana_pregunta.title("Configuración de Muestreo de Auditoría")
    ventana_pregunta.geometry("380x140")
    ventana_pregunta.resizable(False, False)
    ventana_pregunta.transient(root)
    ventana_pregunta.grab_set()
    
    ttk.Label(ventana_pregunta, text="Seleccione el tamaño de muestra para auditar:", font=("Arial", 10, "bold")).pack(pady=12)
    limite_seleccionado = tk.StringVar(value="50")
    
    def seleccionar_opcion(opcion):
        limite_seleccionado.set(opcion)
        ventana_pregunta.destroy()
        
    frame_btn_preg = ttk.Frame(ventana_pregunta)
    frame_btn_preg.pack(pady=5)
    
    ttk.Button(frame_btn_preg, text="Muestra (Primeros 50 reg.)", width=24, command=lambda: seleccionar_opcion("50")).pack(side="left", padx=5)
    ttk.Button(frame_btn_preg, text="Universo Completo", width=18, command=lambda: seleccionar_opcion("completa")).pack(side="left", padx=5)
    
    root.wait_window(ventana_pregunta)
    max_reg = limite_seleccionado.get()
    
    ventana_padre = tk.Toplevel(root)
    txt_modo = "Muestra de Control" if max_reg == "50" else "Carga Completa"
    ventana_padre.title(f"Módulo de Auditoría Estructural y de Atributos WFS ({txt_modo})")
    
    ancho_pantalla = ventana_padre.winfo_screenwidth()
    alto_pantalla = ventana_padre.winfo_screenheight()
    ancho_v = int(ancho_pantalla * 0.92)
    alto_v = int(alto_pantalla * 0.82)
    pos_x = int((ancho_pantalla - ancho_v) / 2)
    pos_y = int((alto_pantalla - alto_v) / 2)
    ventana_padre.geometry(f"{ancho_v}x{alto_v}+{pos_x}+{pos_y}")
    
    notebook_attr = ttk.Notebook(ventana_padre)
    notebook_attr.pack(fill="both", expand=True, padx=10, pady=10)
    
    def tarea_asincrona_atributos_masiva():
        for layer in capas_tildadas:
            layer_name = layer['Nombre']
            status_label.config(text=f"Descargando datos vectoriales para auditoría de {layer_name}...", foreground="blue")
            root.update_idletasks()
            
            request_url = f"{wfs_url}?service=WFS&version=1.0.0&request=GetFeature&typeName={layer_name}&outputFormat=application/json"
            if max_reg == "50":
                request_url += "&maxFeatures=50"
                
            try:
                gdf = gpd.read_file(request_url)
                if gdf is None or gdf.empty:
                    raise Exception("El servidor no retornó registros válidos.")
                
                df_atributos = gdf.drop(columns='geometry', errors='ignore')
                root.after(0, agregar_pestana_tabla_auditoria, notebook_attr, layer_name, df_atributos)
            except Exception as e:
                root.after(0, lambda ln=layer_name, err=e: messagebox.showerror("Fallo de Extracción WFS", f"No se pudo extraer la matriz de {ln}:\n{err}"))
                
        status_label.config(text="Análisis de auditoría completado.", foreground="green")

    threading.Thread(target=tarea_asincrona_atributos_masiva, daemon=True).start()


def agregar_pestana_tabla_auditoria(notebook, layer_name, df_atributos):
    frame_pestana = ttk.Frame(notebook)
    notebook.add(frame_pestana, text=layer_name)
    columnas = list(df_atributos.columns)
    
    columnas_id_ignorar = [c for c in columnas if c.lower() in ['id', 'fid', 'gid', 'objectid', 'uuid', 'pk']]
    columnas_para_duplicados = [c for c in columnas if c not in columnas_id_ignorar]
    if not columnas_para_duplicados: columnas_para_duplicados = columnas
        
    serie_duplicados = df_atributos.duplicated(subset=columnas_para_duplicados, keep=False)
    total_filas_duplicadas = df_atributos.duplicated(subset=columnas_para_duplicados).sum()
    
    total_filas = len(df_atributos)
    total_celdas = df_atributos.size
    celdas_nulas = 0
    columnas_con_nulos = []
    filas_con_vacios = set()
    
    for col in columnas:
        nulos_en_columna = 0
        for idx, val in enumerate(df_atributos[col].values):
            val_str = str(val).strip() 
            if val is None or val_str == "" or val_str.lower() in ["none", "null", "nan"]:
                celdas_nulas += 1
                nulos_en_columna += 1
                filas_con_vacios.add(idx)
        if nulos_en_columna > 0:
            columnas_con_nulos.append(f"{col} ({nulos_en_columna} vacíos)")
            
    porcentaje_completitud = ((total_celdas - celdas_nulas) / total_celdas * 100) if total_celdas > 0 else 0
    resumen_nulos = ", ".join(columnas_con_nulos[:2]) if columnas_con_nulos else "Ninguno"
    if len(columnas_con_nulos) > 2: resumen_nulos += "..."

    frame_diagnostic = tk.LabelFrame(frame_pestana, text=" Diagnóstico de Calidad de Datos ", font=("Arial", 9, "bold"), fg="#1A365D", bg="#F7FAFC")
    frame_diagnostic.pack(fill="x", padx=10, pady=5, side="top")
    
    ttk.Label(frame_diagnostic, text=f"Registros Totales: {total_filas}", font=("Arial", 9, "bold"), background="#F7FAFC").grid(row=0, column=0, padx=15, pady=8, sticky="w")
    color_salud = "#2F855A" if porcentaje_completitud >= 90 else "#C53030"
    tk.Label(frame_diagnostic, text=f"Completitud: {porcentaje_completitud:.1f}%", font=("Arial", 9, "bold"), fg=color_salud, bg="#F7FAFC").grid(row=0, column=1, padx=15, pady=8, sticky="w")
    ttk.Label(frame_diagnostic, text=f"Campos con Vacíos: {resumen_nulos}", font=("Arial", 9, "italic"), background="#F7FAFC").grid(row=0, column=2, padx=15, pady=8, sticky="w")
    color_dup = "#C53030" if total_filas_duplicadas > 0 else "#2F855A"
    tk.Label(frame_diagnostic, text=f"Filas Duplicadas (Omitiendo ID): {total_filas_duplicadas}", font=("Arial", 9, "bold"), fg=color_dup, bg="#F7FAFC").grid(row=0, column=3, padx=15, pady=8, sticky="w")
    
    frame_filtros = ttk.Frame(frame_pestana)
    frame_filtros.pack(fill="x", padx=10, pady=5)
    
    ttk.Label(frame_filtros, text="Buscar en:").pack(side="left", padx=2)
    valores_combo_columnas = ["Todas las columnas"] + columnas
    combo_columnas = ttk.Combobox(frame_filtros, values=valores_combo_columnas, state="readonly", width=22)
    combo_columnas.set("Todas las columnas")
    combo_columnas.pack(side="left", padx=5)
    
    ttk.Label(frame_filtros, text="Texto:").pack(side="left", padx=2)
    entry_busqueda = ttk.Entry(frame_filtros, width=25)
    entry_busqueda.pack(side="left", padx=5)
    modo_filtro_anomalidades = tk.BooleanVar(value=False)
    
    frame_tabla_contenedor = ttk.Frame(frame_pestana)
    frame_tabla_contenedor.pack(fill="both", expand=True, padx=10, pady=5)
    frame_tabla_contenedor.grid_rowconfigure(0, weight=1)
    frame_tabla_contenedor.grid_columnconfigure(0, weight=1)
    
    tree_attr = ttk.Treeview(frame_tabla_contenedor, columns=columnas, show="headings")
    scroll_y = ttk.Scrollbar(frame_tabla_contenedor, orient="vertical", command=tree_attr.yview)
    scroll_x = ttk.Scrollbar(frame_tabla_contenedor, orient="horizontal", command=tree_attr.xview)
    tree_attr.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
    
    for col in columnas:
        tree_attr.heading(col, text=col)
        max_len = len(str(col))
        for val in df_atributos[col].values:
            max_len = max(max_len, len(str(val)))
        ancho_calculado = max(120, min(max_len * 9, 350))
        tree_attr.column(col, width=ancho_calculado, minwidth=90, stretch=False)
        
    tree_attr.grid(row=0, column=0, sticky="nsew")
    scroll_y.grid(row=0, column=1, sticky="ns")
    scroll_x.grid(row=1, column=0, sticky="ew")

    def refrescar_datos_treeview(*args):
        for item in tree_attr.get_children(): tree_attr.delete(item)
        texto_buscado = entry_busqueda.get().lower().strip()
        columna_objetivo = combo_columnas.get()
        solo_anomalos = modo_filtro_anomalidades.get()
        
        for idx, fila in df_atributos.iterrows():
            valores_string = [str(fila[c]) if fila[c] is not None else "" for c in columnas]
            if texto_buscado:
                if columna_objetivo == "Todas las columnas":
                    if not any(texto_buscado in val.lower() for val in valores_string): continue
                else:
                    if texto_buscado not in str(fila[columna_objetivo]).lower(): continue
            if solo_anomalos:
                if not (idx in filas_con_vacios or serie_duplicados.iloc[idx]): continue
            tree_attr.insert("", "end", values=valores_string)

    entry_busqueda.bind("<KeyRelease>", refrescar_datos_treeview)
    combo_columnas.bind("<<ComboboxSelected>>", refrescar_datos_treeview)
    
    def alternar_modo_auditoria():
        if modo_filtro_anomalidades.get():
            modo_filtro_anomalidades.set(False)
            btn_aislar.config(text="🔍 Aislar Anomalías (Vacíos/Duplicados)", style="TButton")
        else:
            modo_filtro_anomalidades.set(True)
            btn_aislar.config(text="👀 Mostrando Solo Errores en Datos", style="Alert.TButton")
        refrescar_datos_treeview()
        
    style.configure("Alert.TButton", font=("Arial", 9, "bold"), foreground="red")
    btn_aislar = ttk.Button(frame_filtros, text="🔍 Aislar Anomalías (Vacíos/Duplicados)", command=alternar_modo_auditoria)
    btn_aislar.pack(side="left", padx=15)
    refrescar_datos_treeview()

    menu_contextual = tk.Menu(tree_attr, tearoff=0)
    menu_contextual.add_command(label="📋 Copiar ID Primario", command=lambda: root.clipboard_append(tree_attr.item(tree_attr.focus(), "values")[0]) if tree_attr.focus() else None)
    tree_attr.bind("<Button-3>", lambda e: menu_contextual.post(e.x_root, e.y_root))

# --- REPORTE LAB GENERATOR CANVAS ---
class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []
    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()
    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            super().showPage()
        super().save()
    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 9)
        self.setFillColor(colors.HexColor("#718096"))
        self.setStrokeColor(colors.HexColor("#E2E8F0"))
        self.setLineWidth(0.5)
        self.line(40, 45, letter[0] - 40, 45)
        self.drawRightString(letter[0] - 40, 30, f"Página {self._pageNumber} de {page_count}")
        self.drawString(40, 30, "Asistente Avanzado Geo-Servicios - Reporte Técnico de Auditoría")
        self.restoreState()

# --- GENERACIÓN DE INFORMES PDF ---
def generar_pdf_auditoria_wms(url_wms, capas, path_guardado):
    doc = SimpleDocTemplate(path_guardado, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=50, bottomMargin=60)
    story = []
    PRIMARY_COLOR, SECONDARY_COLOR = colors.HexColor("#1A365D"), colors.HexColor("#2B6CB0") 
    TEXT_DARK, BG_LIGHT, BORDER_COLOR = colors.HexColor("#2D3748"), colors.HexColor("#F7FAFC"), colors.HexColor("#E2E8F0")    
    ALERT_COLOR, SUCCESS_COLOR = colors.HexColor("#C53030"), colors.HexColor("#2F855A")   
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('MTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=22, textColor=PRIMARY_COLOR)
    subtitle_style = ParagraphStyle('STitle', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=12, textColor=SECONDARY_COLOR, spaceBefore=14, spaceAfter=8)
    body_style = ParagraphStyle('BText', parent=styles['Normal'], fontName='Helvetica', fontSize=10, textColor=TEXT_DARK, spaceAfter=4)
    table_hdr_style = ParagraphStyle('THdr', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, textColor=colors.white)
    table_body_style = ParagraphStyle('TTxt', parent=styles['Normal'], fontName='Helvetica', fontSize=8.5, textColor=TEXT_DARK)
    kpi_title_style = ParagraphStyle('KTitle', fontName='Helvetica-Bold', fontSize=9, textColor=PRIMARY_COLOR, alignment=1)
    kpi_value_style = ParagraphStyle('KValue', fontName='Helvetica-Bold', fontSize=14, textColor=SECONDARY_COLOR, alignment=1)

    story.append(Paragraph("Informe de Auditoría Geo-Servicio WMS", title_style))
    decor_table = Table([[""]], colWidths=[letter[0]-80], rowHeights=[3])
    decor_table.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,-1), SECONDARY_COLOR)]))
    story.append(decor_table)
    story.append(Spacer(1, 15))
    
    story.append(Paragraph(f"<b>Fecha de Evaluación:</b> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", body_style))
    story.append(Paragraph(f"<b>Protocolo Base:</b> OGC WMS Estándar v{WMS_SERVER_METADATA.get('Versión', '1.1.1')}", body_style))
    story.append(Paragraph(f"<b>Latencia de Respuesta:</b> {WMS_SERVER_METADATA.get('Latencia', 'N/A')}", body_style)) 
    story.append(Paragraph(f"<b>Punto de Enlace (URL):</b> <link href='{url_wms}' color='#2B6CB0'><u>{url_wms}</u></link>", body_style))
    story.append(Spacer(1, 10))
    
    story.append(Paragraph("Capacidades Técnicas y Restricciones Legales", subtitle_style))
    formatos = ", ".join(WMS_SERVER_METADATA.get('Formatos', []))
    story.append(Paragraph(f"<b>Formatos de Salida Gráfica:</b> {formatos}", body_style))
    story.append(Paragraph(f"<b>Políticas de Acceso y Restricciones:</b> {WMS_SERVER_METADATA.get('Restricciones', 'Ninguna')}", body_style))
    story.append(Paragraph(f"<b>Costos / Tarifas de Uso:</b> {WMS_SERVER_METADATA.get('Tarifas', 'Gratuito')}", body_style))
    
    capas_sin_resumen = sum(1 for c in capas if not c.get('Resumen', '').strip() or c.get('Resumen').lower() == "sin resumen")
    porcentaje_salud = ((len(capas) - capas_sin_resumen) / len(capas) * 100) if capas else 0
    color_salud = SUCCESS_COLOR if porcentaje_salud >= 75 else ALERT_COLOR
    
    story.append(Paragraph("Métricas de Control", subtitle_style))
    kpi_datos = [
        [Paragraph("Estado de Conexión", kpi_title_style), Paragraph("Capas Publicadas", kpi_title_style), Paragraph("Salud de Metadatos", kpi_title_style)],
        [Paragraph("<font color='#2F855A'>ONLINE</font>", kpi_value_style), Paragraph(str(len(capas)), kpi_value_style), Paragraph(f"<font color='{color_salud.hexval()}'>{porcentaje_salud:.1f}%</font>", kpi_value_style)]
    ]
    kpi_table = Table(kpi_datos, colWidths=[177, 177, 178])
    kpi_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), BG_LIGHT), ('BOX', (0, 0), (-1, -1), 1, BORDER_COLOR), ('INNERGRID', (0, 0), (-1, -1), 0.5, BORDER_COLOR), ('TOPPADDING', (0, 0), (-1, -1), 8), ('BOTTOMPADDING', (0, 0), (-1, -1), 8)]))
    story.append(kpi_table)
    
    story.append(Paragraph("Inventario y Diagnóstico de Capas Disponibles", subtitle_style))
    tabla_datos = [[Paragraph("Nombre Técnico", table_hdr_style), Paragraph("Título Público", table_hdr_style), Paragraph("Sistemas CRS", table_hdr_style), Paragraph("Resumen / Abstract de Capa", table_hdr_style)]]
    
    for c in capas:
        resumen = c.get('Resumen', '').strip()
        if not resumen or resumen.lower() == "sin resumen":
            resumen = f"<font color='{ALERT_COLOR.hexval()}'><i>Falta Metadato.</i></font>"
        tabla_datos.append([Paragraph(c.get('Nombre', 'N/A'), table_body_style), Paragraph(c.get('Título', 'Sin Título'), table_body_style), Paragraph(c.get('CRS', 'EPSG:4326'), table_body_style), Paragraph(resumen, table_body_style)])
    
    tabla_capas = Table(tabla_datos, colWidths=[120, 130, 110, 172], repeatRows=1)
    tabla_capas.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), PRIMARY_COLOR), ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, BG_LIGHT]), ('LINEBELOW', (0, 0), (-1, -1), 0.5, BORDER_COLOR), ('TOPPADDING', (0,0), (-1,-1), 5), ('BOTTOMPADDING', (0,0), (-1,-1), 5)]))
    story.append(tabla_capas)
    doc.build(story, canvasmaker=NumberedCanvas)

def generar_pdf_auditoria_wfs(url_wfs, capas, path_guardado):
    ancho_hoja, alto_hoja = landscape(letter)
    doc = SimpleDocTemplate(path_guardado, pagesize=(ancho_hoja, alto_hoja), rightMargin=40, leftMargin=40, topMargin=50, bottomMargin=60)
    story = []
    
    PRIMARY_COLOR, SECONDARY_COLOR = colors.HexColor("#1A365D"), colors.HexColor("#2B6CB0") 
    TEXT_DARK, BG_LIGHT, BORDER_COLOR = colors.HexColor("#2D3748"), colors.HexColor("#F7FAFC"), colors.HexColor("#E2E8F0")    
    ALERT_COLOR, SUCCESS_COLOR = colors.HexColor("#C53030"), colors.HexColor("#2F855A")   
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('MainTitleWFS', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=22, leading=26, textColor=PRIMARY_COLOR, spaceAfter=4)
    subtitle_style = ParagraphStyle('SubTitleWFS', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=12, leading=16, textColor=SECONDARY_COLOR, spaceBefore=14, spaceAfter=8)
    body_style = ParagraphStyle('BodyTextWFS', parent=styles['Normal'], fontName='Helvetica', fontSize=10, leading=15, textColor=TEXT_DARK, spaceAfter=4)
    table_hdr_style = ParagraphStyle('TableHdrWFS', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, leading=12, textColor=colors.white)
    table_body_style = ParagraphStyle('TableTxtWFS', parent=styles['Normal'], fontName='Helvetica', fontSize=8.5, leading=12, textColor=TEXT_DARK)
    kpi_title_style = ParagraphStyle('KPITitleWFS', fontName='Helvetica-Bold', fontSize=9, leading=11, textColor=PRIMARY_COLOR, alignment=1)
    kpi_value_style = ParagraphStyle('KPIValueWFS', fontName='Helvetica-Bold', fontSize=14, leading=16, textColor=SECONDARY_COLOR, alignment=1)

    story.append(Paragraph("Informe Técnico de Auditoría de Datos Vectoriales (OGC WFS)", title_style))
    decor_table = Table([[""]], colWidths=[ancho_hoja - 80], rowHeights=[3])
    decor_table.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,-1), SECONDARY_COLOR)]))
    story.append(decor_table)
    story.append(Spacer(1, 15))
    
    fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    version_srv = WFS_SERVER_METADATA.get('Version', '1.1.0')
    story.append(Paragraph(f"<b>Institución / Servidor:</b> {WFS_SERVER_METADATA.get('Titulo_Servidor', 'No declarado')}", body_style))
    story.append(Paragraph(f"<b>Fecha de Evaluación:</b> {fecha_actual} | <b>Protocolo Base:</b> OGC WFS v{version_srv} | <b>Latencia:</b> {WFS_SERVER_METADATA.get('Latencia', 'N/A')}", body_style)) 
    story.append(Paragraph(f"<b>Punto de Enlace (URL):</b> <link href='{url_wfs}' color='#2B6CB0'><u>{url_wfs}</u></link>", body_style))
    story.append(Spacer(1, 10))
    
    story.append(Paragraph("Capacidades Legales y de Interoperabilidad", subtitle_style))
    formatos = ", ".join(WFS_SERVER_METADATA.get('Formatos', ['GML', 'GeoJSON']))
    story.append(Paragraph(f"<b>Formatos de Transferencia Admitidos:</b> {formatos}", body_style))
    story.append(Paragraph(f"<b>Políticas de Acceso / Restricciones de Uso:</b> {WFS_SERVER_METADATA.get('Restricciones', 'Ninguna')}", body_style))
    story.append(Paragraph(f"<b>Estructura de Costos / Tarifas:</b> {WFS_SERVER_METADATA.get('Tarifas', 'Gratuito')}", body_style))
    
    total_capas = len(capas)
    col_widths = [140, 160, 70, 90, 280] 
    tabla_datos = [[Paragraph("Nombre Técnico", table_hdr_style), Paragraph("Título Público", table_hdr_style), Paragraph("Geometría", table_hdr_style), Paragraph("Registros", table_hdr_style), Paragraph("Resumen / Descripción", table_hdr_style)]]
    
    capas_sin_resumen = 0
    for c in capas:
        layer_name = c['Nombre']
        resumen = c.get('Resumen', '').strip()
        if not resumen or resumen.lower() in ["sin resumen", "sin resumen declarado en servidor"]:
            capas_sin_resumen += 1
            resumen = f"<font color='{ALERT_COLOR.hexval()}'><i>Falta Metadato descriptivo.</i></font>"
            
        geometria = "Vectorial"
        name_lower = layer_name.lower()
        if any(x in name_lower for x in ['point', 'nodo', 'puntos', 'localidad', 'hitos', 'estacion']): geometria = "Punto"
        elif any(x in name_lower for x in ['line', 'ruta', 'lineas', 'limite', 'ferrocarril']): geometria = "Línea"
        elif any(x in name_lower for x in ['poly', 'muni', 'prov', 'pais', 'area', 'departamento']): geometria = "Polígono"

        tabla_datos.append([Paragraph(layer_name, table_body_style), Paragraph(c.get('Título', 'Sin Título'), table_body_style), Paragraph(geometria, table_body_style), Paragraph(str(c.get('Registros', '0')), table_body_style), Paragraph(resumen, table_body_style)])
    
    porcentaje_salud = ((total_capas - capas_sin_resumen) / total_capas * 100) if total_capas > 0 else 0
    color_salud = SUCCESS_COLOR if porcentaje_salud >= 75 else ALERT_COLOR
    
    story.append(Paragraph("Métricas de Control de Datos", subtitle_style))
    kpi_datos = [
        [Paragraph("Estado de Conexión", kpi_title_style), Paragraph("Capas Auditadas", kpi_title_style), Paragraph("Salud de Metadatos", kpi_title_style)],
        [Paragraph("<font color='#2F855A'>ONLINE</font>", kpi_value_style), Paragraph(str(total_capas), kpi_value_style), Paragraph(f"<font color='{color_salud.hexval()}'>{porcentaje_salud:.1f}%</font>", kpi_value_style)]
    ]
    kpi_table = Table(kpi_datos, colWidths=[242, 242, 244])
    kpi_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), BG_LIGHT), ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('BOX', (0, 0), (-1, -1), 1, BORDER_COLOR), ('INNERGRID', (0, 0), (-1, -1), 0.5, BORDER_COLOR), ('TOPPADDING', (0, 0), (-1, -1), 8), ('BOTTOMPADDING', (0, 0), (-1, -1), 8)]))
    story.append(kpi_table)
    story.append(Spacer(1, 15))
    
    story.append(Paragraph("Matriz de Diagnóstico de Objetos Territoriales", subtitle_style))
    tabla_capas = Table(tabla_datos, colWidths=col_widths, repeatRows=1)
    tabla_capas.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), PRIMARY_COLOR), ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('TOPPADDING', (0, 0), (-1, -1), 5), ('BOTTOMPADDING', (0, 0), (-1, -1), 5), ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, BG_LIGHT]), ('LINEBELOW', (0, 0), (-1, -1), 0.5, BORDER_COLOR)]))
    story.append(tabla_capas)
    
    class NumberedCanvasLandscape(canvas.Canvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved_page_states = []
        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()
        def save(self):
            num_pages = len(self._saved_page_states)
            for state in self._saved_page_states:
                self.__dict__.update(state)
                self.draw_page_number(num_pages)
                super().showPage()
            super().save()
        def draw_page_number(self, page_count):
            self.saveState()
            self.setFont("Helvetica", 9)
            self.setFillColor(colors.HexColor("#718096"))
            self.setStrokeColor(colors.HexColor("#E2E8F0"))
            self.setLineWidth(0.5)
            self.line(40, 45, ancho_hoja - 40, 45) 
            self.drawRightString(ancho_hoja - 40, 30, f"Página {self._pageNumber} de {page_count}")
            self.drawString(40, 30, "Asistente Avanzado Geo-Servicios - Reporte Técnico Vectorial (WFS)")
            self.restoreState()

    doc.build(story, canvasmaker=NumberedCanvasLandscape)

# --- EXCEL FORMATO NATIVO (.xlsx) ---
def aplicar_estilos_excel_servidor(ws, titulo_reporte, meta_datos):
    ws.merge_cells("A1:E1")
    ws["A1"] = titulo_reporte
    ws["A1"].font = Font(name="Segoe UI", size=16, bold=True, color="1A365D")
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 35

    fila = 3
    for k, v in meta_datos.items():
        ws.cell(row=fila, column=1, value=k).font = Font(name="Segoe UI", size=10, bold=True, color="4A5568")
        ws.cell(row=fila, column=2, value=str(v)).font = Font(name="Segoe UI", size=10, color="2D3748")
        ws.row_dimensions[fila].height = 18
        fila += 1
    return fila + 2

def finalizar_formato_tabla_excel(ws, fila_inicio_tabla, columnas_cabecera):
    fill_header = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    border_thin = Side(border_style="thin", color="E2E8F0")
    border_celda = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    ws.row_dimensions[fila_inicio_tabla].height = 25
    for col_idx, texto in enumerate(columnas_cabecera, 1):
        cell = ws.cell(row=fila_inicio_tabla, column=col_idx, value=texto)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center if texto in ["Fila", "Geometría", "Cantidad de Registros", "Sistemas CRS"] else align_left
        cell.border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=Side(border_style="double", color="1A365D"))

    ultima_fila = ws.max_row
    fill_zebra = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")

    for r_idx in range(fila_inicio_tabla + 1, ultima_fila + 1):
        ws.row_dimensions[r_idx].height = 20
        usar_zebra = (r_idx % 2 == 0)
        for c_idx in range(1, len(columnas_cabecera) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = Font(name="Segoe UI", size=10, color="2D3748")
            cell.border = border_celda
            if usar_zebra: cell.fill = fill_zebra
            
            cabecera = columnas_cabecera[c_idx - 1]
            if cabecera in ["Fila", "Geometría", "Sistemas CRS"]: cell.alignment = align_center
            elif cabecera == "Cantidad de Registros":
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0'

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col if cell.row != 1)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 3, 12)

def generar_excel_wms(url_wms, capas, path_guardado):
    if not path_guardado.lower().endswith('.xlsx'): path_guardado = os.path.splitext(path_guardado)[0] + '.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auditoría WMS"
    ws.views.sheetView[0].showGridLines = True

    meta = {
        "Punto de Enlace (URL):": url_wms,
        "Fecha de Evaluación:": datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
        "Protocolo Base:": f"OGC WMS Estándar v{WMS_SERVER_METADATA.get('Versión', '1.1.1')}",
        "Latencia de Servidor:": WMS_SERVER_METADATA.get('Latencia', 'N/A'), 
        "Formatos del Servidor:": ", ".join(WMS_SERVER_METADATA.get('Formatos', []))
    }
    fila_tabla = aplicar_estilos_excel_servidor(ws, "Reporte Técnico de Auditoría Geo-Servicio WMS", meta)
    cabeceras = ["Fila", "Nombre Técnico", "Título Público", "Sistemas CRS", "Resumen / Abstract de Capa"]

    for idx, c in enumerate(capas, 1):
        resumen = c.get('Resumen', '').strip()
        if not resumen or resumen.lower() == "sin resumen": resumen = "Falta Metadato descriptivo."
        fila_actual = fila_tabla + idx
        ws.cell(row=fila_actual, column=1, value=idx)
        ws.cell(row=fila_actual, column=2, value=c.get('Nombre', 'N/A'))
        ws.cell(row=fila_actual, column=3, value=c.get('Título', 'Sin Título'))
        ws.cell(row=fila_actual, column=4, value=c.get('CRS', 'EPSG:4326'))
        ws.cell(row=fila_actual, column=5, value=resumen)

    finalizar_formato_tabla_excel(ws, fila_tabla, cabeceras)
    wb.save(path_guardado)

def generar_excel_wfs(url_wfs, capas, path_guardado):
    if not path_guardado.lower().endswith('.xlsx'): path_guardado = os.path.splitext(path_guardado)[0] + '.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auditoría WFS"
    ws.views.sheetView[0].showGridLines = True

    meta = {
        "Institución / Servidor:": WFS_SERVER_METADATA.get('Titulo_Servidor', 'Servidor WFS'),
        "Punto de Enlace (URL):": url_wfs,
        "Fecha de Evaluación:": datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
        "Latencia de Servidor:": WFS_SERVER_METADATA.get('Latencia', 'N/A'), 
        "Protocolo Base:": f"OGC WFS v{WFS_SERVER_METADATA.get('Version', '1.1.0')}"
    }
    fila_tabla = aplicar_estilos_excel_servidor(ws, "Reporte Técnico de Auditoría de Capas Vectoriales (WFS)", meta)
    cabeceras = ["Nombre Técnico", "Título Público", "Geometría", "Cantidad de Registros", "Resumen / Descripción"]

    for idx, c in enumerate(capas, 1):
        resumen = c.get('Resumen', '').strip()
        if not resumen or resumen.lower() == "sin resumen": resumen = "Falta Metadato descriptivo."
        geometria = "Vectorial"
        name_lower = c.get('Nombre', '').lower()
        if any(x in name_lower for x in ['point', 'nodo', 'puntos', 'localidad', 'hitos', 'estacion']): geometria = "Punto"
        elif any(x in name_lower for x in ['line', 'ruta', 'lineas', 'limite', 'ferrocarril']): geometria = "Línea"
        elif any(x in name_lower for x in ['poly', 'muni', 'prov', 'pais', 'area', 'departamento']): geometria = "Polígono"

        fila_actual = fila_tabla + idx
        ws.cell(row=fila_actual, column=1, value=c.get('Nombre', 'N/A'))
        ws.cell(row=fila_actual, column=2, value=c.get('Título', 'Sin Título'))
        ws.cell(row=fila_actual, column=3, value=geometria)
        try: cant_reg = int(c.get('Registros', 0))
        except: cant_reg = 0
        ws.cell(row=fila_actual, column=4, value=cant_reg)
        ws.cell(row=fila_actual, column=5, value=resumen)

    finalizar_formato_tabla_excel(ws, fila_tabla, cabeceras)
    wb.save(path_guardado)

# --- EXPORTADORES ---
def exportar_informe_wms():
    wms_url = wms_url_entry.get().strip()
    capas_tildadas = [c for c in wms_cached_layers if c.get('Checked')] or wms_cached_layers
    if not capas_tildadas: return
    
    nombre_organismo = "Organismo_No_Registrado"
    for nodo in NODOS_CATALOGO:
        if nodo.get("WMS") and nodo["WMS"].strip() == wms_url:
            nombre_organismo = nodo["Organismo"].replace(" ", "_").replace("/", "-").replace("\\", "-")
            break

    ventana_formato = tk.Toplevel(root)
    ventana_formato.title("Formatos de Informe")
    ventana_formato.geometry("320x180")
    ventana_formato.resizable(False, False)
    ventana_formato.transient(root)
    ventana_formato.grab_set()

    ttk.Label(ventana_formato, text="Seleccione los formatos a exportar:", font=("Arial", 10, "bold")).pack(pady=10)
    var_pdf, var_excel = tk.BooleanVar(value=True), tk.BooleanVar(value=False)
    ttk.Checkbutton(ventana_formato, text="Informe PDF (Diseño Visual)", variable=var_pdf).pack(anchor="w", padx=40, pady=5)
    ttk.Checkbutton(ventana_formato, text="Matriz de Datos Excel (.xlsx)", variable=var_excel).pack(anchor="w", padx=40, pady=5)

    def confirmar_wms():
        if not var_pdf.get() and not var_excel.get(): return
        ventana_formato.destroy()
        carpeta_destino = filedialog.askdirectory(title="Seleccione dónde guardar el informe WMS")
        if not carpeta_destino: return

        nombre_base = f"{nombre_organismo}_{datetime.now().strftime('%Y-%m-%d')}"
        status_label.config(text="Generando informes WMS...", foreground="blue")
        root.update_idletasks()

        archivos_creados = []
        try:
            if var_pdf.get():
                path_pdf = os.path.join(carpeta_destino, nombre_base + ".pdf")
                generar_pdf_auditoria_wms(wms_url, capas_tildadas, path_pdf)
                archivos_creados.append(f"PDF: {path_pdf}")
            if var_excel.get():
                path_excel = os.path.join(carpeta_destino, nombre_base + ".xlsx")
                generar_excel_wms(wms_url, capas_tildadas, path_excel)
                archivos_creados.append(f"Excel: {path_excel}")
            status_label.config(text="Informe WMS guardado.", foreground="green")
            messagebox.showinfo("Éxito", "\n".join(archivos_creados))
        except Exception as e: status_label.config(text=f"Error WMS: {e}", foreground="red")

    ttk.Button(ventana_formato, text="Generar", command=confirmar_wms, width=12).pack(pady=10)

def exportar_informe_wfs():
    wfs_url = wfs_url_entry.get().strip()
    capas_tildadas = [c for c in wfs_cached_layers if c.get('Checked')]
    if not capas_tildadas:
        messagebox.showwarning("Atención", "Por favor, tilde al menos una capa.")
        return
    
    nombre_organismo = "Organismo_No_Registrado"
    for nodo in NODOS_CATALOGO:
        if nodo.get("WFS") and nodo["WFS"].strip() == wfs_url:
            nombre_organismo = nodo["Organismo"].replace(" ", "_").replace("/", "-").replace("\\", "-")
            break

    ventana_formato = tk.Toplevel(root)
    ventana_formato.title("Formatos de Informe")
    ventana_formato.geometry("320x180")
    ventana_formato.resizable(False, False)
    ventana_formato.transient(root)
    ventana_formato.grab_set()

    ttk.Label(ventana_formato, text="Seleccione los formatos a exportar:", font=("Arial", 10, "bold")).pack(pady=10)
    var_pdf, var_excel = tk.BooleanVar(value=True), tk.BooleanVar(value=False)
    ttk.Checkbutton(ventana_formato, text="Informe PDF (Diseño Visual)", variable=var_pdf).pack(anchor="w", padx=40, pady=5)
    ttk.Checkbutton(ventana_formato, text="Matriz de Datos Excel (.xlsx)", variable=var_excel).pack(anchor="w", padx=40, pady=5)

    def confirmar_wfs():
        if not var_pdf.get() and not var_excel.get(): return
        ventana_formato.destroy()
        carpeta_destino = filedialog.askdirectory(title="Seleccione dónde guardar")
        if not carpeta_destino: return

        nombre_base = f"{nombre_organismo}_{datetime.now().strftime('%Y-%m-%d')}"

        def proceso_asincrono_formatos():
            archivos_creados = []
            try:
                version_srv = WFS_SERVER_METADATA.get('Version', '2.0.0')
                for idx, c in enumerate(capas_tildadas, 1):
                    status_label.config(text=f"Auditando capa {idx}/{len(capas_tildadas)}...", foreground="blue")
                    root.update_idletasks()
                    c['Registros'] = obtener_hits_capa(wfs_url, c['Nombre'], version_srv)
                
                if var_pdf.get():
                    path_pdf = os.path.join(carpeta_destino, nombre_base + ".pdf")
                    generar_pdf_auditoria_wfs(wfs_url, capas_tildadas, path_pdf)
                    archivos_creados.append(f"PDF: {path_pdf}")
                if var_excel.get():
                    path_excel = os.path.join(carpeta_destino, nombre_base + ".xlsx")
                    generar_excel_wfs(wfs_url, capas_tildadas, path_excel)
                    archivos_creados.append(f"Excel: {path_excel}")

                actualizar_tabla_wfs_visual()
                status_label.config(text="Informe WFS guardado.", foreground="green")
                messagebox.showinfo("Éxito", "\n".join(archivos_creados))
            except Exception as e: status_label.config(text=f"Error WFS: {e}", foreground="red")

        threading.Thread(target=proceso_asincrono_formatos).start()

    ttk.Button(ventana_formato, text="Generar", command=confirmar_wfs, width=12).pack(pady=10)

# --- MAPAS ---
def configurar_mapas_base(mapa_objeto):
    folium.TileLayer('openstreetmap', name='OpenStreetMap (Color)', show=True).add_to(mapa_objeto)
    folium.TileLayer(tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}', attr='Esri World Imagery', name='Imagen Satelital (Esri)', overlay=False, show=False).add_to(mapa_objeto)
    folium.TileLayer(tiles='https://wms.ign.gob.ar/geoserver/gwc/service/tms/1.0.0/mapabase_gris@EPSG%3A3857@png/{z}/{x}/{y}.png', attr='IGN Argenmap', name='Argenmap Gris (IGN)', tms=True, overlay=False, show=False).add_to(mapa_objeto)

def mostrar_mapa_wms_visual():
    if not wms_cached_layers: return
    capas_tildadas = [c for c in wms_cached_layers if c.get('Checked')]
    if not capas_tildadas: return

    guardar = messagebox.askyesno("Guardar Mapa", "¿Desea guardar el mapa interactivo .html?")
    if guardar:
        path_mapa = filedialog.asksaveasfilename(title="Guardar Visor WMS", filetypes=[("HTML", "*.html")], defaultextension=".html", initialfile="mapa_wms_visual.html")
        if not path_mapa: return
    else:
        path_mapa = f"temp_wms_{uuid.uuid4().hex[:8]}.html"
        ARCHIVOS_TEMPORALES_MAPAS.append(path_mapa)

    try:
        wms_url = wms_url_entry.get().strip()
        centro_mapa, zoom_inicial = [-38.4161, -63.6167], 4
        m = folium.Map(location=centro_mapa, zoom_start=zoom_inicial, tiles=None)
        configurar_mapas_base(m)

        url_limpia = wms_url.split('?')[0]
        for i, row in enumerate(capas_tildadas):
            folium.raster_layers.WmsTileLayer(url=url_limpia, layers=row['Nombre'], name=row['Título'] or row['Nombre'], fmt='image/png', transparent=True, version='1.3.0', overlay=True, show=(i == 0)).add_to(m)

        folium.LayerControl(collapsed=False).add_to(m)
        m.save(path_mapa)
        webbrowser.open(os.path.abspath(path_mapa))
    except Exception as e: status_label.config(text=f"Error WMS: {e}", foreground="red")

def mostrar_mapa_wfs_multi_thread(wfs_url, capas_tildadas, path_mapa):
    try:
        base_url = wfs_url.split('?')[0]
        gdfs_cargados = {}

        for idx, row in enumerate(capas_tildadas, 1):
            name, title = row['Nombre'], row['Título']
            status_label.config(text=f"Mapeando ({idx}/{len(capas_tildadas)}): {name}...", foreground="blue")
            root.update_idletasks()
            
            request_url = f"{base_url}?service=WFS&version=1.0.0&request=GetFeature&typeName={name}&outputFormat=application/json"
            try:
                gdf = gpd.read_file(request_url)
                if gdf is None or gdf.empty: 
                    continue
                if gdf.crs and gdf.crs.to_epsg() != 4326: 
                    gdf = gdf.to_crs(epsg=4326)
                
                # --- SOLUCIÓN AL ERROR DE TIMESTAMP ---
                # Convertimos cualquier columna de tipo fecha/tiempo a texto para que sea compatible con JSON y Folium
                for col in gdf.columns:
                    if gdf[col].dtype == 'datetime64[ns]' or pd.api.types.is_datetime64_any_dtype(gdf[col]):
                        gdf[col] = gdf[col].dt.strftime('%Y-%m-%d %H:%M:%S')
                    elif hasattr(gdf[col], 'dt'): # Por si hay variaciones de zonas horarias (datetimes con tz)
                        try:
                            gdf[col] = gdf[col].astype(str)
                        except:
                            pass

                gdfs_cargados[name] = {"gdf": gdf, "title": title}
            except Exception as e: 
                print(f"Error cargando capa vectorial {name}: {e}")
                pass

        if not gdfs_cargados: 
            status_label.config(text="Error: No se pudieron cargar las geometrías WFS.", foreground="red")
            return

        # Calculamos los límites dinámicos de todas las capas para centrar el mapa automáticamente
        combined_bounds = None
        for data in gdfs_cargados.values():
            bounds = data["gdf"].total_bounds
            if combined_bounds is None:
                combined_bounds = list(bounds)
            else:
                combined_bounds[0] = min(combined_bounds[0], bounds[0])
                combined_bounds[1] = min(combined_bounds[1], bounds[1])
                combined_bounds[2] = max(combined_bounds[2], bounds[2])
                combined_bounds[3] = max(combined_bounds[3], bounds[3])

        m = folium.Map(tiles=None)
        configurar_mapas_base(m)

        colores = ['#ff7800', '#3388ff', '#22ad3c', '#b512b5', '#e02424']
        for i, (layer_name, data) in enumerate(gdfs_cargados.items()):
            gdf_layer = data["gdf"]
            columnas_popup = list(gdf_layer.columns.drop('geometry', errors='ignore'))
            
            folium.GeoJson(
                gdf_layer, 
                name=f"WFS: {data['title']}", 
                popup=folium.GeoJsonPopup(fields=columnas_popup[:10], labels=True), 
                style_function=lambda x, color=colores[i % len(colores)]: {
                    'fillColor': color, 
                    'color': color, 
                    'weight': 2.5
                }
            ).add_to(m)

        folium.LayerControl(collapsed=False).add_to(m)
        m.fit_bounds([[combined_bounds[1], combined_bounds[0]], [combined_bounds[3], combined_bounds[2]]])
        
        m.save(path_mapa)
        webbrowser.open(os.path.abspath(path_mapa))
        status_label.config(text="Mapa WFS desplegado con éxito.", foreground="green")

    except Exception as e: 
        status_label.config(text=f"Error WFS Mapa: {e}", foreground="red")
        
    finally:
        try: progress_bar.pack_forget()
        except: pass
        try: btn_ver_mapa_wfs.config(state="normal")
        except: pass
        try: view_wfs_button.config(state="normal")
        except: pass

def iniciar_visor_wfs():
    wfs_url = wfs_url_entry.get().strip()
    capas_tildadas = [c for c in wfs_cached_layers if c['Checked']]
    if not WFS_SERVER_METADATA or not capas_tildadas: return
        
    guardar = messagebox.askyesno("Guardar Mapa", "¿Desea guardar el mapa .html?")
    if guardar:
        path_mapa = filedialog.asksaveasfilename(title="Guardar Mapa", filetypes=[("HTML", "*.html")], defaultextension=".html")
        if not path_mapa: return
    else:
        path_mapa = f"temp_wfs_{uuid.uuid4().hex[:8]}.html"
        ARCHIVOS_TEMPORALES_MAPAS.append(path_mapa)

    progress_bar.pack(fill="x", padx=10, pady=5)
    progress_bar.config(mode="indeterminate")
    view_wfs_button.config(state="disabled")
    threading.Thread(target=mostrar_mapa_wfs_multi_thread, args=(wfs_url, capas_tildadas, path_mapa)).start()


# =========================================================================
# --- MODIFICACIÓN CLAVE: SOPORTE DE DESCARGA MULTI-CAPA GEOPACKAGE ---
# =========================================================================
def download_wfs_layers_thread(wfs_url, target_path, formato_seleccionado, progress_bar, btn_trigger):
    try:
        capas_a_descargar = [c for c in wfs_cached_layers if c['Checked']]
        base_url = wfs_url.split('?')[0]
        
        # Si es GeoPackage, trabajamos sobre un único archivo unificado
        es_gpkg = (formato_seleccionado == "GeoPackage (.gpkg)")
        
        for i, row in enumerate(capas_a_descargar, 1):
            layer_name = row['Nombre']
            layer_name_cleaned = layer_name.replace(':', '_').replace('/', '_').replace('\\', '_')
            
            status_label.config(text=f"Descargando ({i}/{len(capas_a_descargar)}): {layer_name}...", foreground="blue")
            root.update_idletasks()
            
            # Formatear el formato de salida según capacidades OGC standard
            out_fmt = "application/json"
            if formato_seleccionado == "Shapefile (ZIP)": out_fmt = "SHAPE-ZIP"
            elif formato_seleccionado == "KML": out_fmt = "KML"

            try:
                if es_gpkg:
                    # Para GeoPackage, bajamos como JSON dinámico y Geopandas compila dentro del archivo SQLite (.gpkg)
                    request_url = f"{base_url}?service=WFS&version=1.0.0&request=GetFeature&typeName={layer_name}&outputFormat=application/json"
                    gdf = gpd.read_file(request_url)
                    if gdf is not None and not gdf.empty:
                        # Forzar nombre de capa limpio dentro del GPKG
                        gdf.to_file(target_path, layer=layer_name_cleaned, driver="GPKG")
                else:
                    # Descargas individuales tradicionales (GeoJSON, KML, ZIP)
                    ext = "geojson" if formato_seleccionado == "GeoJSON" else ("kml" if formato_seleccionado == "KML" else "zip")
                    response = requests.get(base_url, params={'service': 'WFS', 'version': '1.0.0', 'request': 'GetFeature', 'typeName': layer_name, 'outputFormat': out_fmt}, timeout=60)
                    if response.status_code == 200:
                        file_out = os.path.join(target_path, f"{layer_name_cleaned}.{ext}")
                        with open(file_out, 'wb') as f:
                            f.write(response.content)
            except Exception as e:
                print(f"Error procesando capa {layer_name}: {e}")
                
            progress_bar['value'] = (i / len(capas_a_descargar)) * 100
            root.update_idletasks()

        status_label.config(text=f"Operación finalizada correctamente.", foreground="green")
        messagebox.showinfo("Éxito", "La descarga y empaquetamiento estructural ha finalizado con éxito.")
    except Exception as e:
        status_label.config(text="Error masivo en la descarga.", foreground="red")
        messagebox.showerror("Error", f"Ocurrió un fallo en el hilo de descarga:\n{e}")
    finally:
        progress_bar.pack_forget()
        btn_trigger.config(state="normal")

def iniciar_descarga_wfs():
    wfs_url = wfs_url_entry.get().strip()
    capas_tildadas = [c for c in wfs_cached_layers if c['Checked']]
    if not wfs_url or not capas_tildadas: 
        messagebox.showwarning("Atención", "Debe tener capas seleccionadas para descargar.")
        return
    
    formato = format_combobox.get()
    
    # Derivación de flujo según el formato espacial
    if formato == "GeoPackage (.gpkg)":
        target_path = filedialog.asksaveasfilename(title="Crear Base de Datos GeoPackage Única", filetypes=[("GeoPackage", "*.gpkg")], defaultextension=".gpkg", initialfile="capas_consolidadas.gpkg")
        if not target_path: return
    else:
        target_path = filedialog.askdirectory(title="Seleccione Carpeta de Destino para las capas individuales")
        if not target_path: return

    progress_bar.pack(fill="x", padx=10, pady=5)
    progress_bar.config(mode="determinate")
    progress_bar['value'] = 0
    download_button.config(state="disabled")
    
    threading.Thread(target=download_wfs_layers_thread, args=(wfs_url, target_path, formato, progress_bar, download_button)).start()


# --- GESTIÓN DE CATÁLOGO ---
def guardar_catalogo():
    try:
        with open("catalogo.json", "w", encoding="utf-8") as f: json.dump(NODOS_CATALOGO, f, indent=4, ensure_ascii=False)
        return True
    except: return False

def agregar_organismo():
    ventana = tk.Toplevel(root)
    ventana.title("Agregar Organismo")
    ventana.geometry("500x220")
    entry_org = ttk.Entry(ventana, width=60)
    entry_wms = ttk.Entry(ventana, width=60)
    entry_wfs = ttk.Entry(ventana, width=60)
    
    ttk.Label(ventana, text="Nombre del Organismo:").pack(anchor="w", padx=10, pady=2)
    entry_org.pack(padx=10, fill="x")
    ttk.Label(ventana, text="URL WMS:").pack(anchor="w", padx=10, pady=2)
    entry_wms.pack(padx=10, fill="x")
    ttk.Label(ventana, text="URL WFS:").pack(anchor="w", padx=10, pady=2)
    entry_wfs.pack(padx=10, fill="x")

    def guardar():
        if not entry_org.get().strip(): return
        NODOS_CATALOGO.append({"Organismo": entry_org.get().strip(), "WMS": entry_wms.get().strip(), "WFS": entry_wfs.get().strip()})
        if guardar_catalogo(): filtrar_lista_catalogo(); ventana.destroy()

    ttk.Button(ventana, text="Guardar", command=guardar).pack(pady=10)

def borrar_organismo(event=None):
    selected = idera_tree.selection()
    if not selected: return
    idx = int(selected[0].replace("nodo_", ""))
    if messagebox.askyesno("Confirmar", "¿Desea eliminar este elemento?"):
        del NODOS_CATALOGO[idx]
        if guardar_catalogo(): filtrar_lista_catalogo()


# --- INTERFAZ GRÁFICA PRINCIPAL ---
root = tk.Tk()
root.title("Asistente Avanzado Geo-Servicios WMS/WFS")
root.geometry("850x680")

style = ttk.Style()
style.configure("TNotebook.Tab", font=("Arial", 9, "bold"))
tab_control = ttk.Notebook(root)
wms_tab, wfs_tab, idera_tab = ttk.Frame(tab_control), ttk.Frame(tab_control), ttk.Frame(tab_control)

tab_control.add(wms_tab, text='Servicios WMS')
tab_control.add(wfs_tab, text='Servicios WFS')
tab_control.add(idera_tab, text='Mi Catálogo Local')
tab_control.pack(expand=True, fill="both", padx=5, pady=5)

# PESTAÑA WMS
ttk.Label(wms_tab, text="URL del servicio WMS:").pack(pady=(10, 0), anchor="w", padx=15)
frame_top_wms = ttk.Frame(wms_tab)
frame_top_wms.pack(pady=5, fill="x", padx=10)
wms_url_entry = ttk.Entry(frame_top_wms, width=60)
wms_url_entry.pack(side="left", expand=True, fill="x", padx=(0,5))
ttk.Button(frame_top_wms, text="Cargar Datos", command=cargar_lista_wms).pack(side="right")

frame_select_actions_wms = ttk.Frame(wms_tab)
frame_select_actions_wms.pack(anchor="w", padx=10, pady=(2, 0))
ttk.Button(frame_select_actions_wms, text="☑ Tildar Todas", command=lambda: cambiar_estado_todas_wms(True), width=15).pack(side="left", padx=2)
ttk.Button(frame_select_actions_wms, text="☐ Destildar Todas", command=lambda: cambiar_estado_todas_wms(False), width=17).pack(side="left", padx=2)

frame_tree_wms = ttk.Frame(wms_tab)
frame_tree_wms.pack(pady=10, padx=10, fill="both", expand=True)
wms_tree = ttk.Treeview(frame_tree_wms, columns=("Sel","Nombre", "Título"), show="headings", height=14)
wms_scroll_y = ttk.Scrollbar(frame_tree_wms, orient="vertical", command=wms_tree.yview)
wms_scroll_x = ttk.Scrollbar(frame_tree_wms, orient="horizontal", command=wms_tree.xview) 
wms_tree.configure(yscrollcommand=wms_scroll_y.set, xscrollcommand=wms_scroll_x.set)
wms_tree.heading("Sel", text="Selección"); wms_tree.heading("Nombre", text="Nombre de la capa"); wms_tree.heading("Título", text="Título de la capa")
wms_tree.column("Sel", width=75, anchor="center"); wms_tree.column("Nombre", width=250); wms_tree.column("Título", width=400)
wms_tree.grid(row=0, column=0, sticky="nsew"); wms_scroll_y.grid(row=0, column=1, sticky="ns"); wms_scroll_x.grid(row=1, column=0, sticky="ew")
frame_tree_wms.grid_rowconfigure(0, weight=1); frame_tree_wms.grid_columnconfigure(0, weight=1)
wms_tree.bind("<Double-1>", alternar_seleccion_item_wms)

frame_bot_wms = ttk.Frame(wms_tab)
frame_bot_wms.pack(pady=5, anchor="e", padx=10)
ttk.Button(frame_bot_wms, text="Generar Informes", command=exportar_informe_wms).pack(side="left", padx=5)
ttk.Button(frame_bot_wms, text="Ver Mapa", command=mostrar_mapa_wms_visual).pack(side="left", padx=5)

# PESTAÑA WFS
ttk.Label(wfs_tab, text="URL del servicio WFS:").pack(pady=(10, 0), anchor="w", padx=15)
frame_top_wfs = ttk.Frame(wfs_tab)
frame_top_wfs.pack(pady=5, fill="x", padx=10)
wfs_url_entry = ttk.Entry(frame_top_wfs, width=60)
wfs_url_entry.pack(side="left", expand=True, fill="x", padx=(0,5))
ttk.Button(frame_top_wfs, text="Cargar Capas WFS", command=cargar_lista_wfs).pack(side="right")

frame_select_actions = ttk.Frame(wfs_tab)
frame_select_actions.pack(anchor="w", padx=10, pady=(2, 0))
ttk.Button(frame_select_actions, text="☑ Tildar Todas", command=lambda: cambiar_estado_todas_wfs(True), width=15).pack(side="left", padx=2)
ttk.Button(frame_select_actions, text="☐ Destildar Todas", command=lambda: cambiar_estado_todas_wfs(False), width=17).pack(side="left", padx=2)

frame_tree_wfs = ttk.Frame(wfs_tab)
frame_tree_wfs.pack(pady=(5, 10), padx=10, fill="both", expand=True)
wfs_tree = ttk.Treeview(frame_tree_wfs, columns=("Sel", "Nombre", "Título"), show="headings", height=12)
wfs_scroll_y = ttk.Scrollbar(frame_tree_wfs, orient="vertical", command=wfs_tree.yview)
wfs_scroll_x = ttk.Scrollbar(frame_tree_wfs, orient="horizontal", command=wfs_tree.xview) 
wfs_tree.configure(yscrollcommand=wfs_scroll_y.set, xscrollcommand=wfs_scroll_x.set)
wfs_tree.heading("Sel", text="Selección"); wfs_tree.heading("Nombre", text="Nombre de la capa"); wfs_tree.heading("Título", text="Título de la capa")
wfs_tree.column("Sel", width=75, anchor="center"); wfs_tree.column("Nombre", width=250); wfs_tree.column("Título", width=400)
wfs_tree.grid(row=0, column=0, sticky="nsew"); wfs_scroll_y.grid(row=0, column=1, sticky="ns"); wfs_scroll_x.grid(row=1, column=0, sticky="ew")
frame_tree_wfs.grid_rowconfigure(0, weight=1); frame_tree_wfs.grid_columnconfigure(0, weight=1)
wfs_tree.bind("<Double-1>", alternar_seleccion_item)

frame_bot_wfs = ttk.Frame(wfs_tab)
frame_bot_wfs.pack(pady=5, fill="x", padx=10)
frame_bot_left = ttk.Frame(frame_bot_wfs)
frame_bot_left.pack(side="left")
ttk.Label(frame_bot_left, text="Formato: ").pack(side="left")

# Agregado GeoPackage a las opciones de descarga
format_combobox = ttk.Combobox(frame_bot_left, values=["GeoPackage (.gpkg)", "GeoJSON", "Shapefile (ZIP)", "KML"], state="readonly", width=18)
format_combobox.set("GeoPackage (.gpkg)")
format_combobox.pack(side="left", padx=2)

download_button = ttk.Button(frame_bot_left, text="Descargar Capa", command=iniciar_descarga_wfs)
download_button.pack(side="left", padx=2)

frame_bot_right = ttk.Frame(frame_bot_wfs)
frame_bot_right.pack(side="right")
ttk.Button(frame_bot_right, text="Ver Tabla de de Atributos", command=ver_tabla_atributos_wfs).pack(side="left", padx=2)
view_wfs_button = ttk.Button(frame_bot_right, text="Ver Mapa", command=iniciar_visor_wfs)
view_wfs_button.pack(side="left", padx=2)
ttk.Button(frame_bot_right, text="Generar Informes", command=exportar_informe_wfs).pack(side="left", padx=2)

# PESTAÑA MI CATÁLOGO
frame_header_idera = ttk.Frame(idera_tab)
frame_header_idera.pack(pady=(10, 5), fill="x", padx=15)
ttk.Label(frame_header_idera, text="Filtrar Organismo: ", font=("Arial", 9, "bold")).pack(side="left")
idera_search_entry = ttk.Entry(frame_header_idera, width=35)
idera_search_entry.pack(side="left", padx=5)
idera_search_entry.bind("<KeyRelease>", filtrar_lista_catalogo)

frame_tree_idera = ttk.Frame(idera_tab)
frame_tree_idera.pack(pady=5, padx=15, fill="both", expand=True)
idera_tree = ttk.Treeview(frame_tree_idera, columns=("Organismo", "WMS_URL", "WFS_URL"), show="headings", height=14)
idera_scroll_y = ttk.Scrollbar(frame_tree_idera, orient="vertical", command=idera_tree.yview)
idera_scroll_x = ttk.Scrollbar(frame_tree_idera, orient="horizontal", command=idera_tree.xview)
idera_tree.configure(yscrollcommand=idera_scroll_y.set, xscrollcommand=idera_scroll_x.set)
idera_tree.heading("Organismo", text="Organismo / Proveedor"); idera_tree.heading("WMS_URL", text="Dirección WMS"); idera_tree.heading("WFS_URL", text="Dirección WFS")
idera_tree.column("Organismo", width=220); idera_tree.column("WMS_URL", width=210); idera_tree.column("WFS_URL", width=210)
idera_tree.grid(row=0, column=0, sticky="nsew"); idera_scroll_y.grid(row=0, column=1, sticky="ns"); idera_scroll_x.grid(row=1, column=0, sticky="ew")
frame_tree_idera.grid_rowconfigure(0, weight=1); frame_tree_idera.grid_columnconfigure(0, weight=1)

frame_bot_idera = ttk.Frame(idera_tab)
frame_bot_idera.pack(pady=10, anchor="center")
ttk.Button(frame_bot_idera, text="Agregar Organismo", command=agregar_organismo, width=20).pack(side="left", padx=5)
ttk.Button(frame_bot_idera, text="Borrar Organismo", command=borrar_organismo, width=18).pack(side="left", padx=5)
ttk.Button(frame_bot_idera, text="Cargar WMS", command=lambda: inyectar_nodo_catalogo('WMS'), width=18).pack(side="left", padx=5)
ttk.Button(frame_bot_idera, text="Cargar WFS", command=lambda: inyectar_nodo_catalogo('WFS'), width=18).pack(side="left", padx=5)

progress_bar = ttk.Progressbar(root, orient="horizontal")
status_label = ttk.Label(root, text="Listo", font=("Arial", 9, "italic"))
status_label.pack(side="bottom", pady=5)

threading.Thread(target=cargar_catalogo_local, daemon=True).start()
root.protocol("WM_DELETE_WINDOW", lambda: [limpiar_archivos_temporales(), root.destroy()])
root.mainloop()